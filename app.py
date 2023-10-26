import openpyxl

def process_cell_text(cell_value):
    """Separate text based on ';' and return text after ';'"""
    if cell_value and ";" in cell_value:
        return cell_value.split(";", 1)[-1].strip()
    return cell_value

def main():
    # Load the workbook and select the active worksheet
    filename = "path_to_your_file.xlsx"
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Get the headers (first row)
    headers = [cell.value.split(";", 1)[0].strip() if cell.value and ";" in cell.value else cell.value
               for cell in sheet[1]]

    # Rename columns with the extracted headers and process each cell
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for idx, cell in enumerate(row, start=1):
            if cell.row == 1:
                sheet.cell(row=cell.row, column=idx, value=headers[idx-1])
            else:
                cell.value = process_cell_text(cell.value)

    # Save the workbook with the "_fixed" suffix
    new_filename = filename.replace(".xlsx", "_fixed.xlsx")
    workbook.save(new_filename)

if __name__ == "__main__":
    main()
